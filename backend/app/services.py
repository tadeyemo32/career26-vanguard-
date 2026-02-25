"""
App services: spreadsheet load, paste parsing, company resolution, email lookup, people search.
Turn almost any data format into structured rows → extract people.
"""
from __future__ import annotations

import csv
import io
import os
import time
import re
from pathlib import Path
from typing import Any, Callable

# Optional Excel
try:
    import openpyxl
    _HAS_EXCEL = True
except ImportError:
    _HAS_EXCEL = False

# Optional PDF
try:
    from pypdf import PdfReader
    _HAS_PDF = True
except ImportError:
    try:
        from PyPDF2 import PdfReader
        _HAS_PDF = True
    except ImportError:
        _HAS_PDF = False


def load_spreadsheet(path: str | Path) -> tuple[list[dict[str, Any]], list[str], str | None]:
    """
    Load CSV or Excel file. Returns (rows as list of dicts, column names, error message or None).
    Rows have keys = column names (from header row).
    """
    path = Path(path)
    if not path.exists():
        return [], [], f"File not found: {path}"
    suffix = path.suffix.lower()
    if suffix == ".csv":
        try:
            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f, skipinitialspace=True)
                cols = reader.fieldnames or []
                rows = list(reader)
            return rows, list(cols), None
        except Exception as e:
            return [], [], str(e)
    if suffix in (".xlsx", ".xls") and _HAS_EXCEL:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                wb.close()
                return [], [], "Empty sheet"
            cols = [str(c).strip() if c is not None else "" for c in header]
            rows = []
            for row in it:
                row = list(row) if row else []
                row = row + [None] * (len(cols) - len(row))
                rows.append(dict(zip(cols, row[: len(cols)])))
            wb.close()
            return rows, cols, None
        except Exception as e:
            return [], [], str(e)
    return [], [], f"Unsupported format: {suffix}. Use .csv or .xlsx"


def load_file_any(path: str | Path) -> tuple[list[dict[str, Any]], list[str], str | None]:
    """
    Load CSV, Excel, or PDF. Returns (rows, column_names, error or None).
    PDF: each line of extracted text becomes a row with key 'company' or 'text'.
    """
    path = Path(path)
    if not path.exists():
        return [], [], f"File not found: {path}"
    suffix = path.suffix.lower()
    if suffix == ".pdf" and _HAS_PDF:
        try:
            reader = PdfReader(path)
            lines = []
            for page in reader.pages:
                text = (page.extract_text() or "").strip()
                for line in text.splitlines():
                    line = line.strip()
                    if line and len(line) > 1:
                        lines.append(line)
            # Build rows: "Name, Company" or single company per line
            rows = []
            for ln in lines:
                if "," in ln:
                    parts = [p.strip() for p in ln.split(",", 1)]
                    if len(parts) == 2:
                        rows.append({"name": parts[0], "company": parts[1]})
                    else:
                        rows.append({"company": ln})
                else:
                    rows.append({"company": ln})
            cols = ["name", "company"] if any("name" in r for r in rows) else ["company"]
            return rows, cols, None
        except Exception as e:
            return [], [], str(e)
    if suffix == ".pdf":
        return [], [], "PDF support requires pypdf or PyPDF2. Install with: pip install pypdf"
    return load_spreadsheet(path)


def extract_companies_from_rows(rows: list[dict[str, Any]]) -> list[str]:
    """
    Extract unique company names from loaded rows. Prefers columns: company, Company,
    company_name, Company Name, organisation; else first column or 'company' key.
    """
    if not rows:
        return []
    candidates = ("company", "Company", "company_name", "Company Name", "organisation", "Organization")
    col = None
    for c in candidates:
        if rows[0].get(c) is not None:
            col = c
            break
    if col is None:
        keys = list(rows[0].keys())
        col = keys[0] if keys else "company"
    seen: set[str] = set()
    out = []
    for r in rows:
        val = (r.get(col) or "")
        if isinstance(val, str):
            val = val.strip()
        else:
            val = str(val).strip() if val is not None else ""
        if val and val not in seen:
            seen.add(val)
            out.append(val)
    return out


def find_people_at_companies(
    companies: list[str],
    job_titles: list[str],
    max_per_company: int = 5,
    find_emails: bool = True,
    delay_seconds: float = 1.2,
) -> list[dict[str, Any]]:
    """
    For each company, search for people with given job titles (e.g. "Director at CompanyName").
    Returns aggregated list of {name, title, company, link, email, confidence}.
    """
    from enrichment.serp import serp_google
    from enrichment.email_finder import find_email_anymail_by_company
    out = []
    titles = [t.strip() for t in job_titles if t.strip()] or ["professional"]
    for company in companies:
        for title in titles[:3]:  # max 3 title variants per company to limit API calls
            q = f'"{title} at {company}" site:linkedin.com/in'
            results = serp_google(q)
            time.sleep(delay_seconds)
            for r in results[:max_per_company]:
                title_str = (r.get("title") or "").strip()
                snippet = (r.get("snippet") or "").strip()
                link = (r.get("link") or "").strip()
                name = ""
                comp = company
                if " - " in title_str:
                    name = title_str.split(" - ")[0].strip()
                    rest = title_str.split(" - ", 1)[1]
                    if " | " in rest:
                        comp = rest.split(" | ")[-1].strip() or company
                    elif " at " in rest:
                        comp = rest.split(" at ", 1)[-1].strip() or company
                else:
                    name = title_str.split("|")[0].strip() if "|" in title_str else title_str
                if not name:
                    continue
                row = {"name": name, "title": title, "company": comp, "link": link, "email": "", "confidence": 0.0}
                if find_emails and name and comp:
                    email, conf = find_email_anymail_by_company(name, comp, timeout=60)
                    row["email"] = email or ""
                    row["confidence"] = conf
                    time.sleep(delay_seconds)
                out.append(row)
    return out


def parse_pasted_data(text: str) -> tuple[list[dict[str, Any]], list[str], str | None]:
    """
    Parse pasted text into rows. Handles:
    - CSV-like (first line = header, comma-separated columns)
    - TSV (tab-separated)
    - One value per line (single column: company or name)
    Returns (rows, column_names, error_message or None).
    """
    text = (text or "").strip()
    if not text:
        return [], [], "No data pasted"
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return [], [], "No lines found"
    # Try CSV: first line has commas and we have multiple lines
    first = lines[0]
    if "," in first and len(lines) >= 1:
        try:
            reader = csv.DictReader(io.StringIO(text), skipinitialspace=True)
            cols = reader.fieldnames or []
            rows = list(reader)
            if rows and cols:
                return rows, list(cols), None
        except Exception:
            pass
    # Try TSV
    if "\t" in first and len(lines) >= 1:
        try:
            reader = csv.DictReader(io.StringIO(text), delimiter="\t", skipinitialspace=True)
            cols = reader.fieldnames or []
            rows = list(reader)
            if rows and cols:
                return rows, list(cols), None
        except Exception:
            pass
    # Single column: one entity per line (company name or "Name, Company")
    rows = []
    cols = ["company"]
    for ln in lines:
        ln = ln.strip()
        if not ln:
            continue
        if "," in ln:
            parts = [p.strip() for p in ln.split(",", 1)]
            if len(parts) == 2:
                rows.append({"name": parts[0], "company": parts[1]})
            else:
                rows.append({"company": ln})
        else:
            rows.append({"company": ln})
    if rows and len(rows[0]) > 1:
        cols = ["name", "company"]
    return rows, cols, None


def resolve_company_to_domain(company_name: str, api_key: str | None = None) -> str:
    """
    Resolve company name to best-guess domain via SerpAPI (e.g. "company name" UK).
    Returns domain or empty string.
    """
    key = api_key or os.environ.get("SERPAPI_KEY")
    if not key or not (company_name or "").strip():
        return ""
    import requests
    q = f'"{company_name.strip()}" UK company'
    url = "https://serpapi.com/search"
    params = {"q": q, "api_key": key, "engine": "google", "num": 5}
    try:
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        data = r.json()
        for o in data.get("organic_results") or []:
            link = (o.get("link") or "").strip()
            if not link or "linkedin.com" in link or "facebook.com" in link:
                continue
            link = link.replace("https://", "").replace("http://", "").split("/")[0]
            link = link.replace("www.", "")
            if link and "." in link and len(link) > 4:
                return link
    except Exception:
        pass
    return ""


def find_email_direct(full_name: str, company: str, use_domain_resolve: bool = False) -> tuple[str | None, float, str]:
    """
    Find email for one person: full_name + company. Uses Anymail (company_name or domain).
    Returns (email, confidence, status_message).
    """
    from enrichment.email_finder import find_email_anymail, find_email_anymail_by_company
    full_name = (full_name or "").strip()
    company = (company or "").strip()
    if not full_name or not company:
        return (None, 0.0, "Name and company required")
    # Prefer domain for accuracy; resolve if requested
    domain = ""
    if use_domain_resolve:
        domain = resolve_company_to_domain(company)
        time.sleep(1.0)
    if domain:
        email, conf = find_email_anymail(full_name, domain, timeout=90)
        return (email, conf, f"domain={domain}" if email else "not found")
    email, conf = find_email_anymail_by_company(full_name, company, timeout=90)
    return (email, conf, "valid" if conf >= 1.0 else "risky" if email else "not found")


def resolve_company_domain_suggestions(company_name: str, api_key: str | None = None, max_suggestions: int = 5) -> list[tuple[str, str]]:
    """
    Web search: resolve company name to possible (display_name, domain) pairs for disambiguation.
    E.g. "Google" -> [("Google (google.com)", "google.com"), ("Google UK", "google.co.uk")].
    """
    key = api_key or os.environ.get("SERPAPI_KEY")
    if not key or not (company_name or "").strip():
        return []
    import requests
    q = f'"{company_name.strip()}" company website'
    url = "https://serpapi.com/search"
    params = {"q": q, "api_key": key, "engine": "google", "num": max_suggestions}
    seen_domains: set[str] = set()
    out: list[tuple[str, str]] = []
    try:
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        data = r.json()
        for o in data.get("organic_results") or []:
            link = (o.get("link") or "").strip()
            title = (o.get("title") or "").strip()
            if not link or "linkedin.com" in link or "facebook.com" in link or "wikipedia.org" in link:
                continue
            domain = link.replace("https://", "").replace("http://", "").split("/")[0].replace("www.", "")
            if not domain or "." not in domain or len(domain) < 4 or domain in seen_domains:
                continue
            seen_domains.add(domain)
            display = f"{title} ({domain})" if title else domain
            out.append((display, domain))
            if len(out) >= max_suggestions:
                break
    except Exception:
        pass
    return out


def find_email_with_log(
    full_name: str,
    company_or_domain: str,
    resolve_domain: bool = True,
) -> tuple[str | None, float, list[str], list[tuple[str, str]]]:
    """
    Find email with a step-by-step log (web-search style). Accepts company name or domain.
    Returns (email, confidence, log_lines, domain_suggestions).
    """
    from enrichment.email_finder import find_email_anymail, find_email_anymail_by_company
    full_name = (full_name or "").strip()
    company_or_domain = (company_or_domain or "").strip()
    log: list[str] = []
    suggestions: list[tuple[str, str]] = []

    if not full_name or not company_or_domain:
        log.append("Name and company/domain required.")
        return (None, 0.0, log, suggestions)

    log.append(f"Input: \"{full_name}\" at \"{company_or_domain}\"")
    # If it looks like a domain (no spaces, has dot), use it directly
    is_domain = " " not in company_or_domain and "." in company_or_domain and len(company_or_domain) > 4
    domain = company_or_domain.replace("www.", "").split("/")[0] if is_domain else ""

    if is_domain:
        log.append(f"Using as domain: {domain}")
    elif resolve_domain:
        log.append("Resolving company to domain (web search)...")
        domain = resolve_company_to_domain(company_or_domain)
        time.sleep(1.0)
        if domain:
            log.append(f"Resolved: {company_or_domain} → {domain}")
            suggestions = resolve_company_domain_suggestions(company_or_domain, max_suggestions=3)
            if suggestions:
                log.append("Other options: " + "; ".join(s[0] for s in suggestions))
        else:
            log.append("No domain found. Trying Anymail with company name.")
    else:
        log.append("Using company name (no domain resolution).")

    if domain:
        log.append(f"Anymail lookup: {full_name} @ {domain}")
        email, conf = find_email_anymail(full_name, domain, timeout=90)
        if email:
            log.append(f"Found: {email} (confidence {conf:.0%})")
            return (email, conf, log, suggestions)
        log.append("No email found for this domain.")
    else:
        log.append(f"Anymail lookup: {full_name} @ {company_or_domain}")
        email, conf = find_email_anymail_by_company(full_name, company_or_domain, timeout=90)
        if email:
            log.append(f"Found: {email} (confidence {conf:.0%})")
            return (email, conf, log, suggestions)
        log.append("No email found.")

    return (None, 0.0, log, suggestions)


def llm_enhance_search_query(user_query: str, api_key: str | None = None) -> str:
    """
    Use LLM to turn a natural-language query into a better LinkedIn-style search query.
    E.g. "IB bankers in London who work for Goldman" -> "investment banker Goldman Sachs London"
    """
    key = api_key or os.environ.get("OPENAI_API_KEY")
    if not key or not (user_query or "").strip():
        return (user_query or "").strip()
    try:
        from openai import OpenAI
        client = OpenAI(api_key=key)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You convert a user's natural language request into a short LinkedIn-style search query to find people. Output only the search query, no explanation. Include job role, company or sector, and location when mentioned. Keep it under 10 words."},
                {"role": "user", "content": user_query.strip()},
            ],
            max_tokens=60,
        )
        text = (response.choices[0].message.content or "").strip()
        return text or user_query.strip()
    except Exception:
        return user_query.strip()


def people_search(
    query: str,
    max_results: int = 20,
    find_emails: bool = True,
    delay_seconds: float = 1.5,
) -> list[dict[str, Any]]:
    """
    Run a people search (e.g. "senior IB bankers London partner"). Uses SerpAPI for
    LinkedIn people results, then optionally finds email via Anymail (company from snippet).
    Returns list of {name, title, company, link, email, confidence}.
    """
    from enrichment.serp import serp_google
    from enrichment.email_finder import find_email_anymail_by_company
    q = f'"{query}" site:linkedin.com/in'
    results = serp_google(q)
    time.sleep(delay_seconds)
    out = []
    for r in results[:max_results]:
        title = (r.get("title") or "").strip()
        snippet = (r.get("snippet") or "").strip()
        link = (r.get("link") or "").strip()
        # Heuristic: "Name - Title | Company" or "Name - Title at Company"
        name = ""
        company = ""
        if " - " in title:
            name = title.split(" - ")[0].strip()
            rest = title.split(" - ", 1)[1]
            if " | " in rest:
                company = rest.split(" | ")[-1].strip()
            elif " at " in rest:
                company = rest.split(" at ", 1)[-1].strip()
            else:
                company = rest.strip()
        else:
            name = title.split("|")[0].strip() if "|" in title else title
        if not name:
            continue
        row = {"name": name, "title": "", "company": company, "link": link, "email": "", "confidence": 0.0}
        if " at " in snippet:
            row["company"] = row["company"] or snippet.split(" at ")[-1].split(".")[0].strip()
        if find_emails and row["company"] and name:
            email, conf = find_email_anymail_by_company(name, row["company"], timeout=60)
            row["email"] = email or ""
            row["confidence"] = conf
            time.sleep(delay_seconds)
        out.append(row)
    return out


def export_results_csv(rows: list[dict[str, Any]], path: str | Path, columns: list[str] | None = None) -> str | None:
    """Export list of dicts to CSV. Returns error message or None."""
    if not rows:
        return "No rows to export"
    path = Path(path)
    cols = columns or list(rows[0].keys())
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow({k: (r.get(k) or "") for k in cols})
        return None
    except Exception as e:
        return str(e)


def export_results_excel(rows: list[dict[str, Any]], path: str | Path, columns: list[str] | None = None) -> str | None:
    """Export to Excel. Returns error message or None."""
    if not _HAS_EXCEL or not rows:
        return "No rows or openpyxl not installed"
    path = Path(path)
    cols = columns or list(rows[0].keys())
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(cols)
        for r in rows:
            ws.append([r.get(c) for c in cols])
        wb.save(path)
        return None
    except Exception as e:
        return str(e)


def save_env_keys(env_path: Path, keys: tuple[str, ...], getter: Callable[[str], str]) -> None:
    """Update .env with API key values from getter(key). Only keys with non-empty values are written. Never log or expose values."""
    existing = env_path.read_text(encoding="utf-8", errors="replace").splitlines() if env_path.exists() else []
    updates = {k: getter(k) for k in keys if getter(k)}
    if not updates:
        return
    out: list[str] = []
    seen: set[str] = set()
    for line in existing:
        m = re.match(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*=(.*)$", line)
        if m and m.group(1) in keys:
            seen.add(m.group(1))
            if m.group(1) in updates:
                out.append(f"{m.group(1)}={updates[m.group(1)]}")
            else:
                out.append(line.rstrip())
        else:
            out.append(line.rstrip())
    for k in keys:
        if k in updates and k not in seen:
            out.append(f"{k}={updates[k]}")
    env_path.write_text("\n".join(out) + "\n", encoding="utf-8")
